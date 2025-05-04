from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def rgb_to_hex(rgb):
    return '{:02X}{:02X}{:02X}'.format(*rgb)

def set_worksheet_dimentions(worksheet, image_width, image_height):
    for i in range(1, image_width + 1):
        col_letter = worksheet.cell(row=1, column=i).column_letter
        worksheet.column_dimensions[col_letter].width = 2

    for i in range(1, image_height + 1):
        worksheet.row_dimensions[i].height = 12

def image_to_excel(image_path, output_excel, resize_to=(100, 100)):
    image = Image.open(image_path).convert('RGB')
    image = image.resize(resize_to)
    width, height = image.size

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Image Pixels"

    for y in range(height):
        for x in range(width):
            r, g, b = image.getpixel((x, y))
            hex_color = rgb_to_hex((r, g, b))
            cell = worksheet.cell(row=y + 1, column=x + 1)
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

    set_worksheet_dimentions(worksheet, width, height)
    workbook.save(output_excel)
    return output_excel
